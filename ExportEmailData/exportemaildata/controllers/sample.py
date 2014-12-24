import openpyxl
workbook = openpyxl.load_workbook(filename = 'D:\Tong\Code\code_python\ExportEmail\ExportEmailData\sample_data\Data5000-1.xlsx', use_iterators = True)
worksheets = workbook.get_sheet_names()

#
#for worksheet in worksheets:
#    print worksheet

worksheet = workbook.get_sheet_by_name('Sheet13')    
print str(worksheet.calculate_dimension());
#print worksheet.columns();
"""

for row in worksheet.iter_rows():
    data = {
        'col01':  row[0].value, # Column A
        'col02': row[1].value, # Column B
        'col03':  row[2].value, # Column C
        'col04':  row[3].value, # Column C
        'col05':  row[4].value, # Column C
        'col06':  row[5].value, # Column C
        'col07':  row[6].value, # Column C
        'col08':  row[7].value, # Column C
        'col09':  row[8].value, # Column C
        'col10':  row[9].value, # Column C
        'col11':  row[10].value, # Column C
        'col12':  row[11].value, # Column C
        'col13':  row[12].value, # Column C
        'col14':  row[13].value, # Column C
        'col15':  row[14].value, # Column C
        'col16':  row[15].value, # Column C
        'col17':  row[16].value, # Column C
        'col18':  row[17].value, # Column C
        'col19':  row[18].value, # Column C
        'col20':  row[19].value, # Column C
        'col21':  row[20].value, # Column C
        'col22':  row[21].value, # Column C
        'col23':  row[22].value, # Column C
        'col24':  row[23].value, # Column C
        'col25':  row[24].value, # Column C
        
    }
    print data;
    
"""
"""
for row in worksheet.iter_rows():
    
    col_d = "";
    for col in row:
        print col.data_type;
    
    #print row[0].value.encode('utf-8');
    
"""

test = {};
test[1] = 555;
test[2] = 666;
print test[1];
print test.get(2);
print test.get(3);


wb = openpyxl.Workbook();
ws = wb.active;

sheet = wb.get_active_sheet();
wb.remove_sheet(sheet);


#sheet = wb.get_active_sheet()

sheet = wb.create_sheet()
sheet.title = "Pi"
 
i =1;
cell_C1 = sheet.cell( row=i, column=1 );
cell_C1.value = "test123";

cell_C2 = sheet.cell( row=i, column=2 );
cell_C2.value = "5555";

path = r'C:\temp\demo.xlsx';
wb.save(path);




print (201%10);
